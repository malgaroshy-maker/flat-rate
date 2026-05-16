"""Parse the POS labor analysis Excel file into typed records."""

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

SHEET_NAME = "POS_ActualLaboursSalesAnalysis"
HEADER_ROW = 10  # 1-indexed
DATA_START_ROW = 11


@dataclass
class LaborRecord:
    branch: str
    wipno: str
    invoice_number: str
    invoice_type: str
    invoice_date: Optional[date]
    code: str  # labor code (2000-4000)
    description: str  # Arabic
    qty: float  # labor hours
    price: float  # hourly rate
    discount_pct: float
    total: float
    account_code: str
    account_name: str
    customer_name: str
    franchise: str  # brand
    model: str
    variant: str
    chassis: str
    reg_date: Optional[date]
    reg_number: str
    department: str  # workshop
    service_advisor: str

    @property
    def effective_total(self) -> float:
        return self.total


def _parse_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _parse_float(value) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _parse_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


_COL_MAP = {
    3: "branch",
    5: "wipno",
    6: "invoice_number",
    7: "invoice_type",
    8: "invoice_date",
    9: "code",
    10: "description",
    11: "qty",
    12: "price",
    14: "discount_pct",
    15: "total",
    16: "account_code",
    17: "account_name",
    19: "customer_name",
    20: "franchise",
    21: "model",
    22: "variant",
    23: "chassis",
    24: "reg_date",
    25: "reg_number",
    26: "department",
    30: "service_advisor",
}


def parse_xlsx(path: str | Path) -> list[LaborRecord]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]
    records: list[LaborRecord] = []

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row_data: dict[str, object] = {}
        for col_idx, field_name in _COL_MAP.items():
            raw = ws.cell(row=row_idx, column=col_idx).value
            if field_name in ("qty", "price", "discount_pct", "total"):
                row_data[field_name] = _parse_float(raw)
            elif field_name in ("invoice_date", "reg_date"):
                row_data[field_name] = _parse_date(raw)
            else:
                row_data[field_name] = _parse_str(raw)

        # Skip empty rows
        if not row_data.get("code") and not row_data.get("description"):
            continue

        records.append(LaborRecord(**{k: row_data.get(k) for k in [
            "branch", "wipno", "invoice_number", "invoice_type",
            "invoice_date", "code", "description", "qty", "price",
            "discount_pct", "total", "account_code", "account_name",
            "customer_name", "franchise", "model", "variant", "chassis",
            "reg_date", "reg_number", "department", "service_advisor",
        ]}))  # type: ignore

    wb.close()
    return records
